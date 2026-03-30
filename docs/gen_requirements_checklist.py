import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "要件定義 確認事項一覧"

# Styles
header_font = Font(name="Meiryo", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
category_font = Font(name="Meiryo", bold=True, size=11)
category_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
normal_font = Font(name="Meiryo", size=10)
priority_high_font = Font(name="Meiryo", size=10, bold=True, color="C00000")
priority_high_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Column widths
col_widths = {"A": 6, "B": 14, "C": 8, "D": 50, "E": 45, "F": 10, "G": 15, "H": 20}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Title
ws.merge_cells("A1:H1")
title_cell = ws["A1"]
title_cell.value = "請求書・納品書メール送信システム — 要件定義 確認事項一覧"
title_cell.font = Font(name="Meiryo", bold=True, size=14)
title_cell.alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 30

# Headers
headers = ["#", "カテゴリ", "ID", "確認事項", "確認が必要な理由", "優先度", "ステータス", "備考・回答"]
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# Data
data = [
    # A. 現行業務フロー
    ("A", "現行業務フロー", [
        ("A-1", "現在の請求書・納品書は誰が、いつ、どのような手順で発行しているか（業務の全体像）", "自動化する範囲と、現行の運用を踏襲すべき箇所を明確にするため", "高"),
        ("A-2", "本システムを実際に操作するのは誰か（営業担当者のみか、経理・事務の方も使うか）", "利用者の役割によって画面設計や操作フローが変わる可能性がある\n※画面モック: https://html-samples.s3.ap-northeast-1.amazonaws.com/tokaisteel/01_select-department.html", "高"),
        ("A-3", "営業担当者の人数（全体および1部門あたり）", "画面の担当者フィルター機能や、表示データ量の見積もりに必要", "中"),
        ("A-4", "1日/1部門あたりの請求書・納品書の発行件数（通常時・ピーク時）", "一括送信処理の時間見積もりおよびシステム容量の設計に必要", "高"),
        ("A-5", "請求書と納品書は同じ伝票に対して両方発行されるのか、それぞれ別の伝票か", "データの分類方法および伝票番号の管理方法に影響するため", "高"),
        ("A-6", "締め日・発行タイミングのサイクル（例：月末締め翌月発行）", "請求書・納品書がどのタイミングで発生するかを把握し、システムの処理フローに反映するため", "中"),
    ]),
    # B. 軽技Web
    ("B", "軽技Web", [
        ("B-1", "【依頼】軽技Webの開発用ログインID・パスワードの発行", "システム開発の着手に必要。お客様にて発行をお願いしたい", "最高"),
        ("B-2", "軽技Webに接続元のIP制限やVPN等のアクセス制限はあるか", "クラウド環境からアクセスできるかどうかで、システム構成とコストが大きく変わる", "最高"),
    ]),
    # C. 電子帳簿保存法（電帳法）対応
    ("C", "電子帳簿保存法対応", [
        ("C-1", "改ざん防止措置の方針（事務処理規程の策定 or システムで訂正・削除履歴を管理）", "電子帳簿保存法で改ざん防止が義務。規程のみなら開発不要、システム対応なら履歴管理機能の追加が必要", "高"),
        ("C-2", "既存の電子帳簿保存法への対応状況（他システムで対応済みの部分はあるか）", "本システムでの対応範囲を確定するため。既に対応済みであれば重複を避ける", "高"),
        ("C-3", "PDF控えの保存期間の方針（法定7年 or 安全策で10年）", "クラウドストレージの保持設定やコストに直結する", "高"),
        ("C-4", "送信済み書類の検索機能（日付・金額の範囲指定、取引先名、組み合わせ検索）の追加について", "電子帳簿保存法では7年分のデータを上記の条件で検索できることが必要。現在の設計（過去1週間表示のみ）から変更が発生する。当面1部門のみの運用であれば、お見積もりの部門選択機能の工数を検索機能に振り替えることで追加費用なしで対応できる可能性がある", "最高"),
    ]),
    # D. 送信ルール・運用
    ("D", "送信ルール・運用", [
        ("D-1", "メール送信ON/OFFのデフォルト値は何を基準に決めるか（得意先の属性、契約内容等）", "得意先マスタの備考2フィールドの運用ルールを決めるため", "高"),
        ("D-2", "備考2の入力フォーマットの取り決め（お客様側で整備予定）。請求書と納品書で個別にON/OFFを設定する必要はあるか", "個別設定の有無によって備考2のフォーマットが変わる。フォーマット確定後にお客様側でデータを整備していただく流れ", "高"),
        ("D-3", "1つの得意先に複数のメールアドレスを設定する必要はあるか。また、すべて宛先（To）でよいか（CC/BCCの区別が必要か）", "複数アドレスやCC/BCC対応が必要な場合、画面やデータの設計が複雑になる", "中"),
        ("D-4", "得意先のメールアドレスの管理方法。備考2に送信可否と合わせて入れるか、基幹側に送信先マスタを新設してもらえるか", "メールアドレスの格納先によって軽技Webからの取得方法やデータ整備の進め方が変わる", "高"),
        ("D-5", "送信元メールアドレスは全社共通か、部門ごとに分けるか", "メール配信サービスの設定やメール文面のテンプレートに影響する", "中"),
        ("D-6", "メールの一括送信は毎日何時が望ましいか。部門や曜日で変えたいか。また、送信時間の変更は画面からの操作は不要（変更時は弊社にて対応）でよいか", "送信時間を画面から変更できるようにすると開発コストが増加する", "中"),
        ("D-7", "メール送信が失敗した場合のリカバリ運用（画面から手動で再送する機能は必要か）", "再送機能が必要な場合、画面に「再送」ボタンの追加が必要", "中"),
        ("D-8", "送信済みの書類について、取り消しや再送信を行う運用はあるか", "必要であれば差し戻し機能の追加開発が発生する", "中"),
        ("D-9", "定時の一括送信以外に、その場で即時送信する機能は必要か", "急ぎの請求書を翌日まで待てないケースがあるか。必要であれば「今すぐ送信」機能の追加開発が発生する", "高"),
        ("D-10", "本システムを介さずに手動でメール送信した場合の扱いをどうするか", "手動で送信すると本システムに送信記録が残らず、履歴の一元管理が崩れる。電子帳簿保存法の観点や二重送信の防止にも影響する", "高"),
        ("D-11", "メール本文や帳票に、得意先ごとの備考・コメント等を添えたいケースはあるか", "必要であれば確定時に備考欄を入力できる機能の追加や、メールテンプレート・PDFレイアウトの変更が発生する", "中"),
        ("D-12", "得意先からメールに返信があった場合の対応方針。返信先は送信元アドレスと同じでよいか、部門や担当者ごとに分けるか、または返信不可（no-reply）とするか", "返信を受け付ける場合、本文なしで返信されるとメールアドレスからしか送信元を特定できない点に注意（件名に伝票番号・得意先名を含めることで軽減可能）。運用負荷を避けるならno-replyも選択肢", "中"),
        ("D-13", "メールの件名・本文のテンプレートは誰が決めるか（お客様で文案をご用意いただくか、弊社で叩き台を作成するか）", "第1回の段階で進め方を決めておくと、以降の確認がスムーズに進む", "中"),
        ("D-14", "土日祝日もメールを送信するか（営業日のみか、カレンダーに関係なく毎日か）", "営業日のみの場合、祝日カレンダーの管理方法を決める必要がある", "中"),
    ]),
    # E. 帳票（PDF）
    ("E", "帳票（PDF）", [
        ("E-1", "【依頼】現行の紙帳票（請求書・納品書）の現物またはPDFサンプルのご提供", "PDFテンプレート設計の基礎資料として使用。インボイス対応状況や請求書と納品書のレイアウトの違いもあわせて確認する", "高"),
        ("E-2", "【依頼】会社ロゴ・社印・角印等の画像データのご提供", "PDF帳票への組み込みに使用する", "中"),
        ("E-3", "1枚に収まらないほど明細行が多い伝票は実際にあるか（1伝票あたりの最大明細数の目安）", "ある場合は複数ページへの改ページ処理の設計が必要になる。ない場合は1ページ固定で簡素化できる", "中"),
    ]),
    # F. アクセス制御・セキュリティ
    ("F", "アクセス制御・セキュリティ", [
        ("F-1", "【依頼】本システムを利用するオフィスの固定IPアドレス", "システムへのアクセス制限の設定に必須", "最高"),
        ("F-2", "リモートワークや外出先からシステムにアクセスする必要はあるか", "IP制限のみで運用可能か、VPN等の追加対策が必要かの判断材料", "高"),
        ("F-3", "複数の拠点からアクセスする必要はあるか", "拠点ごとにIPアドレスの登録が必要になる", "中"),
        ("F-4", "営業担当者ごとのアクセス制限は不要でよいか（全員が全部門の操作を行える前提）", "不要であればログイン機能を省略でき、開発コストを抑えられる", "中"),
    ]),
    # G. インフラ・環境
    ("G", "インフラ・環境", [
        ("G-1", "軽技Webへのクラウド環境からの接続方法の確認。①IP許可リストへの追加は可能か ②VPN接続の発行は可能か ③その他の接続方法はあるか", "開発の最大のブロッカー候補。接続方法によりシステム構成と月額コストが大きく変わる（例：IP許可方式の場合 約4,800円/月の追加費用が発生）", "最高"),
        ("G-2", "メール配信サービス（SendGrid）の契約状況（既存アカウントの有無、送信元ドメインのDNS設定権限）", "メール送信機能の前提条件。未契約の場合は新規契約が必要", "高"),
        ("G-5", "SendGridの利用料金の費用負担はお客様か弊社か", "契約主体と費用負担を明確にしておく必要がある", "中"),
        ("G-3", "【依頼】AWSアカウントの作成（お客様名義で作成、弊社にて構築・運用）", "お客様にアカウントを作成いただき、弊社に操作権限を付与いただく。AWSの利用料金はお客様に直接発生する", "高"),
        ("G-4", "独自ドメインでのシステム公開は必要か（例：invoice.tokai-steel.co.jp）", "必要な場合、ドメイン取得やSSL証明書の設定作業が追加で発生する", "中"),
    ]),
    # H. 運用・保守
    ("H", "運用・保守", [
        ("H-1", "エラー通知や送信結果サマリーの送信先メールアドレス", "システムからの通知メールの宛先を決めるため", "中"),
        ("H-2", "軽技WebのUI変更時の保守対応について", "軽技Webの画面が変更された場合、データ取得機能の修正が必要になる。対応の範囲・費用について事前に認識を合わせておきたい", "高"),
        ("H-3", "過去データの保持期間（画面上の表示は過去1週間だが、それ以外のアーカイブは必要か）", "電子帳簿保存法で最低7年の保存が必要（C-3と連動）。データベース・ファイルストレージ両方の保持設定に影響する", "高"),
        ("H-4", "テスト運用時のメール誤送信防止策（テスト用アドレスへの限定送信、段階的な対象拡大等）", "本番の得意先に誤ってテストメールが届くことを防ぐため、テスト運用の進め方を事前に合意しておく", "中"),
    ]),
]

row = 4
seq = 1
for cat_code, cat_name, items in data:
    # Category header row
    for col_idx in range(1, 9):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = category_fill
        cell.font = category_font
        cell.border = thin_border
        cell.alignment = wrap_alignment
    ws.cell(row=row, column=2, value=f"{cat_code}. {cat_name}")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    row += 1

    for item_id, question, reason, priority in items:
        ws.cell(row=row, column=1, value=seq).font = normal_font
        ws.cell(row=row, column=2, value=f"{cat_code}. {cat_name}").font = normal_font
        ws.cell(row=row, column=3, value=item_id).font = normal_font
        ws.cell(row=row, column=4, value=question).font = normal_font
        ws.cell(row=row, column=5, value=reason).font = normal_font
        ws.cell(row=row, column=6, value=priority).font = normal_font
        status_cell = ws.cell(row=row, column=7, value="未確認")
        status_cell.font = normal_font
        ws.cell(row=row, column=8, value="").font = normal_font

        # Apply border & alignment to all cells
        for col_idx in range(1, 9):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = thin_border
            cell.alignment = wrap_alignment

        # Highlight blocker items
        if priority == "最高":
            for col_idx in range(1, 9):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = priority_high_fill
            ws.cell(row=row, column=6).font = priority_high_font

        ws.row_dimensions[row].height = 40
        row += 1
        seq += 1

# Data validation for status column
from openpyxl.worksheet.datavalidation import DataValidation
status_dv = DataValidation(
    type="list",
    formula1='"未確認,確認済み,保留,不要"',
    allow_blank=True
)
status_dv.error = "リストから選択してください"
status_dv.errorTitle = "入力エラー"
ws.add_data_validation(status_dv)
status_dv.add(f"G4:G{row - 1}")

# Freeze pane
ws.freeze_panes = "A4"

# Auto filter
ws.auto_filter.ref = f"A3:H{row - 1}"

# Save
output_path = "/Users/sumou-no-oujisama/Dropbox/koo/project/東海鋼材工業様/docs/要件定義_確認事項一覧.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")